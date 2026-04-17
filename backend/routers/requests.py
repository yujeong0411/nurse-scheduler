"""근무신청 저장·조회·현황·엑셀 export/import"""
import io
import os
import re
import sys
import tempfile
from datetime import datetime
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from ..database import get_db, db_nurses, get_period_by_id
from ..deps import get_current_admin, get_current_any
from ..schemas import RequestOut, RequestsUpsertBody, SubmissionStatus, NurseScoreOut, AssignmentLogEntry
from ..config import settings

router = APIRouter(prefix="/requests", tags=["근무신청"])


def _row_to_out(row: dict) -> RequestOut:
    return RequestOut(
        id=row["id"],
        nurse_id=row["nurse_id"],
        day=row["day"],
        code=row["code"],
        is_or=row.get("is_or", False),
        note=row.get("note") or '',
        submitted_at=str(row.get("submitted_at", "")),
        condition=row.get("condition", "B"),
        score=row.get("score", 100),
    )


@router.get("/{period_id}", response_model=list[RequestOut])
def get_all_requests(period_id: str, _: dict = Depends(get_current_admin)):
    """전체 간호사 신청 조회 (관리자)"""
    db = get_db()
    res = db.table("requests").select("*").eq("period_id", period_id).execute()
    return [_row_to_out(r) for r in res.data]


@router.get("/{period_id}/me", response_model=list[RequestOut])
def get_my_requests(period_id: str, current: dict = Depends(get_current_any)):
    """본인 신청만 조회 (간호사)"""
    db = get_db()
    nurse_id = current["sub"]
    res = (
        db.table("requests")
        .select("*")
        .eq("period_id", period_id)
        .eq("nurse_id", nurse_id)
        .execute()
    )
    return [_row_to_out(r) for r in res.data]


@router.get("/{period_id}/status", response_model=list[SubmissionStatus])
def get_submission_status(period_id: str, _: dict = Depends(get_current_admin)):
    """각 간호사 제출 여부 + 제출 시각"""
    db = get_db()
    nurses = db_nurses(db).order("sort_order").execute().data

    # 제출한 간호사 목록 (최신 submitted_at 기준)
    req_res = (
        db.table("requests")
        .select("*")
        .eq("period_id", period_id)
        .order("submitted_at", desc=True)
        .execute()
    )
    submitted_map: dict[str, str] = {}
    for r in req_res.data:
        nid = r["nurse_id"]
        if nid not in submitted_map:
            submitted_map[nid] = str(r.get("submitted_at", ""))

    return [
        SubmissionStatus(
            nurse_id=n["id"],
            name=n["name"],
            submitted_at=submitted_map.get(n["id"]),
        )
        for n in nurses
    ]


@router.put("/{period_id}/{nurse_id}", response_model=list[RequestOut])
def upsert_requests(
    period_id: str,
    nurse_id: str,
    body: RequestsUpsertBody,
    current: dict = Depends(get_current_any),
):
    """간호사 본인 또는 관리자가 신청 일괄 저장"""
    # 권한 확인: nurse는 본인만
    if current["role"] == "nurse" and current["sub"] != nurse_id:
        raise HTTPException(403, "본인 신청만 수정할 수 있습니다.")

    db = get_db()
    # period 유효성
    period = get_period_by_id(db, period_id)
    if not period:
        raise HTTPException(404, "해당 기간을 찾을 수 없습니다.")

    # 마감 시간 체크 (간호사만)
    if current["role"] == "nurse" and period.get("deadline"):
        dl_str = period["deadline"]
        try:
            dl_dt = datetime.fromisoformat(dl_str if "T" in dl_str else dl_str + "T23:59")
            if datetime.now() > dl_dt:
                raise HTTPException(403, "신청 마감이 지났습니다.")
        except HTTPException:
            raise
        except Exception:
            pass

    now_iso = datetime.utcnow().isoformat()

    # ── 우선순위: 병가 제외 대상 코드 ──
    _SKIP_PRIORITY = {"병가", "법휴", "필수"}

    # A조건 최대 3개 검증 (병가 제외)
    new_a_count = sum(
        1 for item in body.items
        if item.condition == 'A' and item.code not in _SKIP_PRIORITY
    )
    if new_a_count > 3:
        raise HTTPException(400, f"A조건은 월 최대 3개까지 신청 가능합니다. (현재 {new_a_count}개)")

    # 점수 = 100 - 현재 신청 항목 전체 차감액 (매 저장마다 처음부터 재계산)
    new_score = 100 - sum(
        (1 if item.condition == 'A' else 3)
        for item in body.items
        if item.code not in _SKIP_PRIORITY
    )

    # nurse_scores upsert
    db.table("nurse_scores").upsert({
        "period_id": period_id,
        "nurse_id": nurse_id,
        "score": new_score,
    }, on_conflict="period_id,nurse_id").execute()

    # 기존 신청 전체 삭제 후 재삽입
    db.table("requests").delete().eq("period_id", period_id).eq("nurse_id", nurse_id).execute()

    if not body.items:
        return []

    rows = [
        {
            "period_id": period_id,
            "nurse_id": nurse_id,
            "day": item.day,
            "code": item.code,
            "is_or": item.is_or,
            "note": item.note or '',
            "submitted_at": now_iso,
            "condition": item.condition if item.code not in _SKIP_PRIORITY else 'B',
            "score": new_score,  # 차감 후 점수를 스냅샷으로 저장
        }
        for item in body.items
    ]
    res = db.table("requests").insert(rows).execute()
    return [_row_to_out(r) for r in res.data]


@router.get("/{period_id}/score/{nurse_id}", response_model=NurseScoreOut)
def get_nurse_score(period_id: str, nurse_id: str, current: dict = Depends(get_current_any)):
    """간호사 본인 점수 조회"""
    if current["role"] == "nurse" and current["sub"] != nurse_id:
        raise HTTPException(403)
    db = get_db()
    res = (
        db.table("nurse_scores")
        .select("score")
        .eq("period_id", period_id)
        .eq("nurse_id", nurse_id)
        .execute()
    )
    score = res.data[0]["score"] if res.data else 100
    return NurseScoreOut(nurse_id=nurse_id, score=score)


@router.get("/{period_id}/scores", response_model=list[NurseScoreOut])
def get_all_scores(period_id: str, _: dict = Depends(get_current_admin)):
    """해당 기간 전체 간호사 점수 일괄 조회 (관리자 전용)"""
    db = get_db()
    res = (
        db.table("nurse_scores")
        .select("nurse_id, score")
        .eq("period_id", period_id)
        .execute()
    )
    return [NurseScoreOut(nurse_id=r["nurse_id"], score=r["score"]) for r in res.data]


@router.post("/{period_id}/reset-scores")
def reset_scores(period_id: str, _: dict = Depends(get_current_admin)):
    """해당 기간 모든 간호사 점수 100으로 초기화 (관리자 전용)"""
    db = get_db()
    nurses = db_nurses(db).execute().data
    rows = [{"period_id": period_id, "nurse_id": n["id"], "score": 100} for n in nurses]
    db.table("nurse_scores").upsert(rows, on_conflict="period_id,nurse_id").execute()
    return {"reset": len(rows)}


@router.post("/{period_id}/recalc-scores")
def recalc_scores(period_id: str, _: dict = Depends(get_current_admin)):
    """기존 신청 데이터 기준으로 전체 점수 재계산 (관리자 전용)"""
    _SKIP_PRIORITY = {"병가", "법휴", "필수"}
    db = get_db()
    nurses = db_nurses(db).execute().data
    req_res = db.table("requests").select("nurse_id, code, condition").eq("period_id", period_id).execute()

    # nurse_id별 차감액 합산
    deductions: dict[str, int] = {}
    for r in req_res.data:
        if r["code"] in _SKIP_PRIORITY:
            continue
        nid = r["nurse_id"]
        deductions[nid] = deductions.get(nid, 0) + (1 if r.get("condition") == "A" else 3)

    rows = [
        {"period_id": period_id, "nurse_id": n["id"], "score": 100 - deductions.get(n["id"], 0)}
        for n in nurses
    ]
    db.table("nurse_scores").upsert(rows, on_conflict="period_id,nurse_id").execute()
    return {"recalculated": len(rows)}


@router.get("/{period_id}/assignment-log/{day}/{code}", response_model=list[AssignmentLogEntry])
def get_assignment_log(period_id: str, day: int, code: str, _: dict = Depends(get_current_admin)):
    """날짜-근무 단위 배정 근거 조회 (관리자 전용)"""
    db = get_db()
    res = (
        db.table("assignment_log")
        .select("*")
        .eq("period_id", period_id)
        .eq("day", day)
        .eq("code", code)
        .order("rank")
        .execute()
    )
    nurses = {n["id"]: n["name"] for n in db_nurses(db).execute().data}
    return [
        AssignmentLogEntry(
            nurse_id=r["nurse_id"],
            name=nurses.get(r["nurse_id"], ""),
            code=r.get("code", ""),
            requested_codes=r.get("requested_codes") or r.get("code", ""),
            condition=r["condition"],
            score=r["score"],
            rank=r["rank"],
            is_random=r["is_random"],
            is_assigned=r["is_assigned"],
        )
        for r in res.data
    ]


@router.get("/{period_id}/export")
def export_requests_excel(period_id: str, _: dict = Depends(get_current_admin)):
    """신청현황 xlsx 다운로드 — request_example 포맷"""
    db = get_db()
    period = get_period_by_id(db, period_id)
    if not period:
        raise HTTPException(404)

    nurses = db_nurses(db).order("sort_order").execute().data
    req_res = db.table("requests").select("*").eq("period_id", period_id).execute().data

    # 부서명 조회
    try:
        dept_res = db.table("departments").select("name").eq("id", settings.department_id).single().execute()
        dept_name = dept_res.data.get("name", "") if dept_res.data else ""
    except Exception:
        dept_name = ""

    # nurse_id → {day: {codes: [...], note: str}} 맵
    req_map: dict[str, dict[int, dict]] = {n["id"]: {} for n in nurses}
    for r in req_res:
        nid = r["nurse_id"]
        day = r["day"]
        code = r["code"]
        is_or = r.get("is_or", False)
        note = r.get("note") or ""
        if day not in req_map.setdefault(nid, {}):
            req_map[nid][day] = {"codes": [], "note": note}
        req_map[nid][day]["codes"].append((code, is_or))
        if note:
            req_map[nid][day]["note"] = note

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import date, timedelta

    start = date.fromisoformat(period["start_date"])
    end = start + timedelta(days=27)
    WD = ["월", "화", "수", "목", "금", "토", "일"]
    NUM_DAYS = 28

    wb = Workbook()
    ws = wb.active
    ws.title = "근무신청현황"

    YELLOW_FILL = PatternFill("solid", fgColor="fcfb92")
    WEEKEND_FILL = PatternFill("solid", fgColor="F2F2F2")
    HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
    TITLE_FILL = PatternFill("solid", fgColor="4472C4")
    CENTER = Alignment(horizontal="center", vertical="center")
    BLACK_BORDER = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    def apply_border(cell):
        cell.border = BLACK_BORDER

    # ── 행1: 타이틀 (병합)
    total_cols = NUM_DAYS + 1  # 이름 열 + 날짜 열
    title_text = f"{start.strftime('%Y.%m.%d')} ~ {end.strftime('%Y.%m.%d')}  {dept_name} 근무신청표"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(1, 1, title_text)
    title_cell.font = Font(bold=True, size=12, color="FFFFFF")
    title_cell.fill = TITLE_FILL
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 22

    # ── 행2: 날짜
    c0 = ws.cell(2, 1, "")
    c0.fill = HEADER_FILL; apply_border(c0)
    for i in range(NUM_DAYS):
        d = start + timedelta(days=i)
        wd = d.weekday()
        cell = ws.cell(2, i + 2, f"{d.day}일")
        cell.alignment = CENTER
        cell.font = Font(bold=True, size=9, color="CC0000" if wd >= 5 else "000000")
        cell.fill = WEEKEND_FILL if wd >= 5 else HEADER_FILL
        apply_border(cell)

    # ── 행3: 이름 헤더 + 요일
    c1 = ws.cell(3, 1, "이름")
    c1.font = Font(bold=True, size=10); c1.alignment = CENTER; c1.fill = HEADER_FILL; apply_border(c1)
    for i in range(NUM_DAYS):
        d = start + timedelta(days=i)
        wd = d.weekday()
        cell = ws.cell(3, i + 2, WD[wd])
        cell.alignment = CENTER
        cell.font = Font(size=9, color="CC0000" if wd >= 5 else "333333")
        cell.fill = WEEKEND_FILL if wd >= 5 else HEADER_FILL
        apply_border(cell)

    # ── 행4+: 간호사별
    for nurse in nurses:
        row_idx = ws.max_row + 1
        shifts_raw = req_map.get(nurse["id"], {})

        nc = ws.cell(row_idx, 1, nurse["name"])
        nc.font = Font(size=10); nc.alignment = CENTER; apply_border(nc)

        for i in range(NUM_DAYS):
            day = i + 1
            d = start + timedelta(days=i)
            wd = d.weekday()
            col = i + 2
            cell = ws.cell(row_idx, col)

            fixed_off = nurse.get("fixed_weekly_off")
            is_fixed_off = fixed_off is not None and fixed_off != "" and int(fixed_off) == wd

            if is_fixed_off:
                cell.value = "주"
                cell.fill = WEEKEND_FILL if wd >= 5 else PatternFill("solid", fgColor="FFFFFF")
                cell.alignment = CENTER
                cell.font = Font(size=9, color="666666")
            elif shifts_raw.get(day):
                entry = shifts_raw[day]
                codes = entry["codes"]
                note = entry.get("note", "")
                or_entries = [c for c, is_or in codes if is_or]
                non_or = [c for c, is_or in codes if not is_or]
                if or_entries:
                    cell.value = "/".join(or_entries)
                else:
                    cell.value = non_or[0] if non_or else ""
                cell.fill = YELLOW_FILL
                cell.alignment = CENTER
                if note:
                    from openpyxl.comments import Comment
                    cell.comment = Comment(note, "간호사")
            else:
                cell.value = ""
                if wd >= 5:
                    cell.fill = WEEKEND_FILL
                cell.alignment = CENTER
            apply_border(cell)

    # ── 열 너비
    ws.column_dimensions["A"].width = 12
    for i in range(NUM_DAYS):
        ws.column_dimensions[get_column_letter(i + 2)].width = 6

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    filename = f"{start.strftime('%Y.%m.%d')}~{end.strftime('%Y.%m.%d')}_신청표.xlsx"
    encoded = quote(filename, safe='')
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


@router.post("/{period_id}/import")
def import_requests_excel(
    period_id: str,
    file: UploadFile = File(...),
    _: dict = Depends(get_current_admin),
):
    """신청현황 xlsx 가져오기 — export와 동일한 포맷"""
    from openpyxl import load_workbook

    db = get_db()
    period = get_period_by_id(db, period_id)
    if not period:
        raise HTTPException(404, "기간을 찾을 수 없습니다.")

    nurses = db_nurses(db).order("sort_order").execute().data
    name_to_id = {n["name"].strip(): n["id"] for n in nurses}

    content = file.file.read()
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(content)
    tmp.close()

    try:
        wb = load_workbook(tmp.name, read_only=True, data_only=True)
        ws = wb.active

        # 헤더 행 찾기 (날짜 숫자 20개 이상인 행)
        header_row = None
        day_cols: dict[int, int] = {}  # day_index(1-28) → col_number
        for row in ws.iter_rows():
            temp_cols = []
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value).strip()
                raw = val[:-1].strip() if val.endswith("일") else val
                try:
                    d = int(raw)
                    if 1 <= d <= 31:
                        temp_cols.append((cell.column, d))
                except ValueError:
                    pass
            if len(temp_cols) >= 20:
                header_row = row[0].row
                temp_cols.sort(key=lambda x: x[0])
                day_cols = {i + 1: col for i, (col, _) in enumerate(temp_cols)}
                break

        _CODE_MAP = {
            "VAC": "휴가", "휴": "휴가",
            "병": "병가", "공": "공가", "경": "경가", "생": "생휴",
            "법": "법휴", "오프": "OFF", "주휴": "주",
        }
        _VALID_CODES = {
            "D", "E", "N", "중2", "D9", "D1", "중1",
            "OFF", "POFF", "주", "법휴", "수면", "생휴", "휴가", "병가",
            "특휴", "공가", "경가", "보수", "필수", "번표",
        }

        def _normalize(c: str) -> str:
            if "수면" in c:
                return "수면"
            for s in ["D", "E", "N"]:
                if c.replace(" ", "") == f"{s}제외":
                    return f"{s} 제외"
            upper = c.upper()
            if upper in _VALID_CODES:
                return upper
            return _CODE_MAP.get(upper, _CODE_MAP.get(c, c))

        if not header_row or not day_cols:
            wb.close()
            raise HTTPException(400, "날짜 헤더를 찾을 수 없습니다. 올바른 신청현황 파일인지 확인해주세요.")

        # 이름 열 찾기 ("이름" 셀 위치)
        min_day_col = min(day_cols.values())
        name_col = 1
        for search_row in ws.iter_rows(min_row=max(1, header_row - 1), max_row=header_row + 2,
                                        max_col=min_day_col - 1):
            for cell in search_row:
                if cell.value and str(cell.value).strip() == "이름":
                    name_col = cell.column

        # 데이터 시작 행 찾기 (요일행 건너뜀)
        data_start = header_row + 1
        for check_row in ws.iter_rows(min_row=header_row + 1, max_row=min(header_row + 3, ws.max_row),
                                       max_col=min_day_col + 6):
            vals = {str(c.value).strip() for c in check_row if c.value}
            if vals & {"이름", "일", "월", "화", "수", "목", "금", "토"}:
                data_start = check_row[0].row + 1
                break

        # 데이터 파싱
        now_iso = datetime.utcnow().isoformat()
        imported, skipped = 0, []

        for row in ws.iter_rows(min_row=data_start):
            row_dict = {cell.column: cell.value for cell in row}
            name = str(row_dict.get(name_col) or "").strip()
            if not name:
                continue
            nurse_id = name_to_id.get(name)
            if not nurse_id:
                skipped.append(name)
                continue

            items = []
            for day_idx, col in day_cols.items():
                val = row_dict.get(col)
                if not val:
                    continue
                raw = str(val).strip()
                # "공가(예비군)" 형태 → code="공가", note="예비군"
                _m = re.match(r'^([^(/]+)\(([^)]*)\)', raw)
                note = _m.group(2).strip() if _m else ""
                code = _m.group(1).strip() if _m else raw
                if not code or code == "주":  # 주휴는 자동 처리라 skip
                    continue
                # "/" 구분자 OR 신청 처리
                if "/" in code:
                    for part in code.split("/"):
                        part = _normalize(part.strip())
                        if part:
                            items.append({
                                "period_id": period_id, "nurse_id": nurse_id,
                                "day": day_idx, "code": part,
                                "is_or": True, "note": note, "submitted_at": now_iso,
                                "condition": "B", "score": 100,
                            })
                else:
                    items.append({
                        "period_id": period_id, "nurse_id": nurse_id,
                        "day": day_idx, "code": _normalize(code),
                        "is_or": False, "note": note, "submitted_at": now_iso,
                        "condition": "B", "score": 100,
                    })

            db.table("requests").delete().eq("period_id", period_id).eq("nurse_id", nurse_id).execute()
            if items:
                db.table("requests").insert(items).execute()
            imported += 1

        wb.close()
    finally:
        try:
            os.unlink(tmp.name)
        except Exception:
            pass

    msg = f"{imported}명 신청 가져오기 완료"
    if skipped:
        msg += f" (미매칭 {len(skipped)}명: {', '.join(skipped[:5])}{'...' if len(skipped) > 5 else ''})"
    return {"imported": imported, "skipped": skipped, "message": msg}
