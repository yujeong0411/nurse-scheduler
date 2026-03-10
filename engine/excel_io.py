"""엑셀 가져오기 / 내보내기 — 응급실 간호사 근무표

가져오기:
  - 근무신청표 (달력 격자): 간호사 이름 + 요청사항
  - 근무표 규칙: 간호사 속성 (역할, 직급, 특수조건)

내보내기:
  - 근무표 시트 (17종 근무/휴무 컬러)
  - 통계 시트
"""
import calendar
import io
import re
from datetime import date, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from engine.models import (
    Nurse, Request, Rules, Schedule,
    WORK_SHIFTS, OFF_TYPES, ALL_CODES,
)


class EncryptedFileError(Exception):
    """파일이 암호화(비밀번호 보호)되어 있습니다."""


def load_workbook_safe(filepath: str, password: str | None = None, **kw):
    """암호화 여부를 감지하고, 필요하면 복호화 후 workbook 반환.

    - 암호화 없음: 일반 load_workbook 호출
    - 암호화 + password 있음: msoffcrypto로 복호화 후 BytesIO에서 열기
    - 암호화 + password 없음: EncryptedFileError 발생
    """
    try:
        import msoffcrypto
    except ImportError:
        return load_workbook(filepath, **kw)

    with open(filepath, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        if not office_file.is_encrypted():
            return load_workbook(filepath, **kw)
        if password is None:
            raise EncryptedFileError("파일이 암호화되어 있습니다. 비밀번호가 필요합니다.")
        office_file.load_key(password=password)
        buf = io.BytesIO()
        office_file.decrypt(buf)
    buf.seek(0)
    return load_workbook(buf, **kw)


# ══════════════════════════════════════════
# 스타일 정의
# ══════════════════════════════════════════

# 근무별 색상
FILLS = {
    # "D":   PatternFill(start_color="DAF0F3", fill_type="solid"),
    # "D9":  PatternFill(start_color="B4E1E8", fill_type="solid"),  # 중간 계열
    # "D1":  PatternFill(start_color="B4E1E8", fill_type="solid"),  # 중간 계열
    # "중1":  PatternFill(start_color="FDEBD0", fill_type="solid"), # 중간 계열
    # "중2":  PatternFill(start_color="FDEBD0", fill_type="solid"), # 중간 계열
    # "E":   PatternFill(start_color="FDE9D9", fill_type="solid"),
    # "N":   PatternFill(start_color="E4DFEC", fill_type="solid"),
    "OFF": PatternFill(start_color="fcfb92", fill_type="solid"),
    "주":   PatternFill(start_color="fcfb92", fill_type="solid"),
    "법휴": PatternFill(start_color="fcfb92", fill_type="solid"),
    "생휴":   PatternFill(start_color="fcfb92", fill_type="solid"),
    "수면": PatternFill(start_color="fcfb92", fill_type="solid"),
    "POFF": PatternFill(start_color="fcfb92", fill_type="solid"),
    "휴가":   PatternFill(start_color="fcfb92", fill_type="solid"),
    "병가":   PatternFill(start_color="fcfb92", fill_type="solid"),
    "특휴": PatternFill(start_color="fcfb92", fill_type="solid"),
    "공가":   PatternFill(start_color="fcfb92", fill_type="solid"),
    "경가":   PatternFill(start_color="fcfb92", fill_type="solid"),
    "보수":   PatternFill(start_color="fcfb92", fill_type="solid"),
    "필수":   PatternFill(start_color="fcfb92", fill_type="solid"),
    "번표":   PatternFill(start_color="fcfb92", fill_type="solid"),
}
FONTS = {
    "D":   Font(size=10),
    # "D9":  Font(color="2E75B6", bold=True, size=9),  # 중간 계열
    # "D1":  Font(color="2E75B6", bold=True, size=9),  # 중간 계열
    # "중1":  Font(color="BF8F00", bold=True, size=9),  # 중간 계열
    "중2":  Font(size=10),
    "E":   Font(size=10),
    "N":   Font(color="d61506", size=10),
    "OFF": Font(size=10),
    "주":   Font(size=10),
    "법휴": Font(size=10),
    "생휴":   Font(size=10),
    "수면": Font(size=10),
    "POFF": Font(size=10),
    "휴가": Font(size=10),
    "병가": Font(size=10),
    "특휴": Font(size=10),
    "공가": Font(size=10),
    "경가": Font(size=10),
    "보수": Font(size=10),
    "필수": Font(size=10),
    "번표": Font(size=10),
}

HEADER_FILL = PatternFill(start_color="013976", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
WEEKEND_FILL = PatternFill(start_color="F2F2F2", fill_type="solid")
SHORTAGE_FILL = PatternFill(start_color="FFCCCC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
RED_BORDER = Border(
    left=Side(style="medium", color="FF0000"),
    right=Side(style="medium", color="FF0000"),
    top=Side(style="medium", color="FF0000"),
    bottom=Side(style="medium", color="FF0000"),
)
CENTER = Alignment(horizontal="center", vertical="center")



# ══════════════════════════════════════════
# 내보내기
# ══════════════════════════════════════════

def export_schedule(schedule: Schedule, rules: Rules, filepath: str):
    """근무표 + 통계를 엑셀로 내보내기"""
    wb = Workbook()

    # ── Sheet 1: 근무표 ──
    ws = wb.active
    ws.title = "근무표"

    start_date = schedule.start_date
    end_date = schedule.date_of(28)
    num_days = schedule.num_days
    nurses = schedule.nurses
    weekday_names = ["월", "화", "수", "목", "금", "토", "일"]

    # 타이틀
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_days + 8)
    title = f"{start_date.strftime('%Y.%m.%d')} ~ {end_date.strftime('%Y.%m.%d')} 응급실 근무표"
    ws.cell(1, 1, title)
    ws.cell(1, 1).font = Font(bold=True, size=14, color="013976")
    ws.cell(1, 1).alignment = Alignment(horizontal="center")

    # 헤더 (행3)
    stat_cols = ["D", "중2", "E", "N", "OFF", "총 근무", "주말", "휴가잔여", "생휴", "잔여수면",]
    headers = ["이름"] + [f"{d}" for d in range(1, num_days + 1)] + stat_cols
    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # 요일 행 (행4)
    ws.cell(4, 1, "요일")
    ws.cell(4, 1).font = Font(size=10)
    ws.cell(4, 1).alignment = CENTER
    for d in range(1, num_days + 1):
        wd = schedule.weekday_index(d)
        cell = ws.cell(4, d + 1, weekday_names[wd])
        cell.alignment = CENTER
        cell.font = Font(size=9, color="CC0000" if wd >= 5 else "333333")
        if wd >= 5:
            cell.fill = WEEKEND_FILL

    # 요청사항 조회 맵 구성
    _off_set = set(OFF_TYPES)
    req_map: dict[tuple[int, int], list[str]] = {}
    is_or_map: dict[tuple[int, int], bool] = {}
    for r in schedule.requests:
        key = (r.nurse_id, r.day)
        if r.is_or:
            req_map.setdefault(key, []).append(r.code)
            is_or_map[key] = True
        else:
            req_map[key] = [r.code]
            is_or_map[key] = False

    # 간호사별 데이터
    weekend_days = [d for d in range(1, num_days + 1) if schedule.is_weekend(d)]

    for i, nurse in enumerate(nurses):
        row = 5 + i
        ws.cell(row, 1, nurse.name)
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 1).alignment = CENTER
        ws.cell(row, 1).border = THIN_BORDER

        d_cnt, 중2_cnt, e_cnt, n_cnt, off_cnt = 0, 0, 0, 0, 0

        for d in range(1, num_days + 1):
            shift = schedule.get_shift(nurse.id, d)
            cell = ws.cell(row, d + 1, shift)
            cell.alignment = CENTER

            # 요청사항 매칭 체크
            key = (nurse.id, d)
            req_codes = req_map.get(key, [])
            is_violation = False
            is_matched = False
            req_display = ""

            if req_codes:
                is_or = is_or_map.get(key, False)
                req_display = "/".join(req_codes) if is_or else req_codes[0]

                if any("제외" in c for c in req_codes):
                    for c in req_codes:
                        if "제외" in c:
                            banned = c.split()[0]
                            if shift == banned:
                                is_violation = True
                                break
                    if not is_violation:
                        is_matched = True
                else:
                    for c in req_codes:
                        if c in _off_set and shift in _off_set:
                            is_matched = True
                            break
                        if c == shift:
                            is_matched = True
                            break
                    if not is_matched:
                        is_violation = True

            # 배경/테두리: 매칭 → 노란색, 불일치 → 흰색+빨간테두리, 없음 → 흰색(주말 연회색)
            if is_matched:
                cell.fill = PatternFill(start_color="FFFF66", fill_type="solid")
                cell.border = THIN_BORDER
            elif is_violation:
                cell.fill = PatternFill(start_color="FFFFFF", fill_type="solid")
                cell.border = RED_BORDER
                cell.comment = Comment(f"요청: {req_display}", "시스템")
            else:
                if schedule.weekday_index(d) >= 5:
                    cell.fill = WEEKEND_FILL
                else:
                    cell.fill = PatternFill(start_color="FFFFFF", fill_type="solid")
                cell.border = THIN_BORDER

            if shift in FONTS:
                cell.font = FONTS[shift]

            if shift == "D":
                d_cnt += 1
            elif shift == "중2":
                중2_cnt += 1
            elif shift == "E":
                e_cnt += 1
            elif shift == "N":
                n_cnt += 1
            elif shift in OFF_TYPES or shift == "OFF":
                off_cnt += 1

        total_work = d_cnt + 중2_cnt + e_cnt + n_cnt
        wk_work = sum(
            1 for d in weekend_days
            if schedule.get_shift(nurse.id, d) in WORK_SHIFTS
        )

        # 휴가잔여/생휴/잔여수면 계산
        vac_used = sum(1 for d in range(1, num_days + 1)
                       if schedule.get_shift(nurse.id, d) == "휴가")
        vac_remain = nurse.vacation_days - vac_used
        menst_cnt = sum(1 for d in range(1, num_days + 1)
                        if schedule.get_shift(nurse.id, d) == "생휴")
        sleep_cnt = sum(1 for d in range(1, num_days + 1)
                        if schedule.get_shift(nurse.id, d) == "수면")
        sleep_earned = (1 if n_cnt >= rules.sleep_N_monthly else 0) + (1 if nurse.pending_sleep else 0)
        sleep_remain = sleep_earned - sleep_cnt

        stat_vals = [d_cnt, 중2_cnt, e_cnt, n_cnt, off_cnt, total_work, wk_work,
                     vac_remain, menst_cnt if menst_cnt else "", sleep_remain if sleep_remain > 0 else ""]
        for j, val in enumerate(stat_vals):
            cell = ws.cell(row, num_days + 2 + j, val)
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            cell.font = Font(size=10)

    # 집계 행
    sep_row = 5 + len(nurses)
    for si, shift_type in enumerate(["D", "중2", "E", "N"]):
        agg_row = sep_row + 1 + si
        ws.cell(agg_row, 1, f"{shift_type} 인원")
        ws.cell(agg_row, 1).font = Font(size=10)
        ws.cell(agg_row, 1).alignment = CENTER

        for d in range(1, num_days + 1):
            count = sum(
                1 for n in nurses
                if schedule.get_shift(n.id, d) == shift_type
            )
            cell = ws.cell(agg_row, d + 1, count)
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            cell.font = Font(size=10)

            min_req = rules.get_daily_staff(shift_type)
            if count < min_req:
                cell.fill = SHORTAGE_FILL
                cell.font = Font(size=10, color="CC0000")

    # 컬럼 너비
    ws.column_dimensions["A"].width = 12
    for d in range(1, num_days + 1):
        ws.column_dimensions[get_column_letter(d + 1)].width = 6
    for j in range(len(stat_cols)):
        ws.column_dimensions[get_column_letter(num_days + 2 + j)].width = 7

    # ── Sheet 2: 통계 ──
    ws2 = wb.create_sheet("통계")
    ws2.cell(1, 1, f"{start_date.strftime('%Y.%m.%d')} ~ {end_date.strftime('%Y.%m.%d')} 개인별 통계")
    ws2.cell(1, 1).font = Font(bold=True, size=14, color="013976")

    stat_headers = ["이름", "직급", "역할", "D", "중2", "E", "N",
                    "OFF", "총근무", "N비율", "주말근무", "휴가잔여", "생휴", "잔여수면"]
    for c, h in enumerate(stat_headers, 1):
        cell = ws2.cell(3, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    for i, nurse in enumerate(nurses):
        row = 4 + i
        d_cnt = sum(
            1 for d in range(1, num_days + 1)
            if schedule.get_shift(nurse.id, d) == "D"
        )
        중2_cnt = sum(
            1 for d in range(1, num_days + 1)
            if schedule.get_shift(nurse.id, d) == "중2"
        )
        e_cnt = sum(
            1 for d in range(1, num_days + 1)
            if schedule.get_shift(nurse.id, d) == "E"
        )
        n_cnt = schedule.get_day_count(nurse.id, "N")
        off_cnt = num_days - d_cnt - 중2_cnt - e_cnt - n_cnt
        total_work = d_cnt + 중2_cnt + e_cnt + n_cnt
        n_ratio = f"{n_cnt / total_work * 100:.0f}%" if total_work > 0 else "0%"
        wk_work = sum(
            1 for d in weekend_days
            if schedule.get_shift(nurse.id, d) in WORK_SHIFTS
        )

        # 휴가잔여/생휴/수면 계산
        vac_used = sum(1 for d in range(1, num_days + 1)
                       if schedule.get_shift(nurse.id, d) == "휴가")
        vac_remain = nurse.vacation_days - vac_used
        menst_cnt = sum(1 for d in range(1, num_days + 1)
                        if schedule.get_shift(nurse.id, d) == "생휴")
        sleep_cnt = sum(1 for d in range(1, num_days + 1)
                        if schedule.get_shift(nurse.id, d) == "수면")
        sleep_earned = (1 if n_cnt >= rules.sleep_N_monthly else 0) + (1 if nurse.pending_sleep else 0)
        sleep_remain = sleep_earned - sleep_cnt

        data = [nurse.name, nurse.grade or "일반", nurse.role or "-",
                d_cnt, 중2_cnt, e_cnt, n_cnt, off_cnt, total_work, n_ratio, wk_work,
                vac_remain, menst_cnt if menst_cnt else "", sleep_remain if sleep_remain > 0 else ""]
        for c, val in enumerate(data, 1):
            cell = ws2.cell(row, c, val)
            cell.alignment = CENTER
            cell.border = THIN_BORDER

    for c in range(1, len(stat_headers) + 1):
        ws2.column_dimensions[get_column_letter(c)].width = 10

    wb.save(filepath)


# ══════════════════════════════════════════
# 가져오기: 근무 규칙 엑셀 → 간호사 속성
# ══════════════════════════════════════════

def import_nurse_rules(filepath: str, password: str | None = None) -> list[Nurse]:
    """근무표_규칙.xlsx에서 간호사 목록 + 속성 불러오기

    형식:
      C열: 이름
      D열: 비고1 (역할) — 책임만, 외상, 혼자 관찰, 급성구역, 준급성, 격리구역(소아)
      E열: 비고2 (직급) — 책임, 서브차지
      F열: 비고3 (특수) — 임산부, 남자
      G열: 비고4 (근무형태) — 주4일제
    """
    wb = load_workbook_safe(filepath, password, read_only=True, data_only=True)
    ws = wb.active

    # 헤더 행 찾기 ("이름" 포함 행)
    header_row = None
    name_col = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            val = str(cell.value).strip() if cell.value else ""
            if val == "이름":
                header_row = cell.row
                name_col = cell.column
                break
        if header_row:
            break

    if not header_row or not name_col:
        wb.close()
        return []

    # 비고 컬럼 위치 (이름 기준 상대)
    # 일반적으로 D=비고1, E=비고2, F=비고3, G=비고4
    col_role = name_col + 1    # 비고1
    col_grade = name_col + 2   # 비고2
    col_special = name_col + 3 # 비고3
    col_work = name_col + 4    # 비고4

    nurses = []
    nurse_id = 1

    for row in ws.iter_rows(min_row=header_row + 1):
        # 이름
        name_cell = row[name_col - 1] if name_col - 1 < len(row) else None
        name = str(name_cell.value).strip() if name_cell and name_cell.value else ""
        if not name:
            continue

        nurse = Nurse(id=nurse_id, name=name)

        # 비고1: 역할
        if col_role - 1 < len(row):
            val = str(row[col_role - 1].value).strip() if row[col_role - 1].value else ""
            if val:
                nurse.role = val

        # 비고2: 직급
        if col_grade - 1 < len(row):
            val = str(row[col_grade - 1].value).strip() if row[col_grade - 1].value else ""
            if val in ("책임", "서브차치", "서브차지"):
                nurse.grade = "서브차지" if "서브" in val else val

        # 비고3: 특수
        if col_special - 1 < len(row):
            val = str(row[col_special - 1].value).strip() if row[col_special - 1].value else ""
            if "임산부" in val:
                nurse.is_pregnant = True
            if "남자" in val or "남" == val:
                nurse.is_male = True

        # 비고4: 근무형태
        if col_work - 1 < len(row):
            val = str(row[col_work - 1].value).strip() if row[col_work - 1].value else ""
            if "주4" in val or "4일" in val:
                nurse.is_4day_week = True

        nurses.append(nurse)
        nurse_id += 1

    wb.close()
    return nurses


# ══════════════════════════════════════════
# 가져오기: 근무신청표 → 요청사항
# ══════════════════════════════════════════

def _normalize_code(val: str) -> str | None:
    """엑셀 셀 값을 표준 코드로 변환

    Returns: 표준 코드 또는 None (무시할 값)
    """
    val = val.strip()
    if not val:
        return None

    # 슬래시 복합 코드 (D/VAC, N/OFF 등) → import_requests에서 OR 처리
    if "/" in val:
        return None

    # 대소문자 통일
    upper = val.upper()

    # 제외 요청
    for s in ["D", "E", "N"]:
        no_space = val.replace(" ", "")
        if no_space == f"{s}제외":
            return f"{s} 제외"

    # 괄호 포함 시 괄호 앞 부분만 추출 (예: "공가(예비군)" → "공가")
    if "(" in val:
        val = val[:val.index("(")].strip()
        upper = val.upper()
        if not val:
            return None

    # 정확한 매칭
    exact_map = {
        "D": "D", "E": "E", "N": "N",
        "D9": "D", "D1": "D",
        "중1": "중1", "중2": "중2",
        "OFF": "OFF", "오프": "OFF",
        "주": "주", "주휴": "주",
        "법": "법휴", "법휴": "법휴",
        "생휴": "생휴", "생": "생휴",
        "수면": "수면",
        "VAC": "휴가", "휴가": "휴가", "휴": "휴가",
        "병가": "병가", "병": "병가",
        "공가": "공가", "공": "공가",
        "경가": "경가", "경": "경가",
        "특휴": "특휴",
        "보수": "보수",
        "필수": "필수",
        "번표": "번표",
        "POFF": "POFF",
    }

    if upper in exact_map:
        return exact_map[upper]
    if val in exact_map:
        return exact_map[val]

    # "수면" 포함 시 → "수면" (예: "3월수면", "1,2월수면", "수면(2월)")
    if "수면" in val:
        return "수면"

    # off 소문자
    if upper == "OFF":
        return "OFF"

    return None


def _find_day_columns(ws) -> tuple[int, dict[int, int], int, int] | None:
    """날짜 헤더 행, 날짜 열 매핑, 이름 열, 데이터 시작 행을 탐색

    열 순서대로 스케줄 day 1, 2, 3, ... 으로 매핑.
    4주(28일) 스케줄이 1일 시작이 아닐 수 있음 (예: 29일, 30일, 31일, 1일, ..., 25일).

    Returns:
        (header_row, day_cols, name_col, data_start) 또는 None (탐색 실패)
        - header_row: 날짜 헤더가 있는 행 번호
        - day_cols: {schedule_day(1-based): column_index} 열 순서 기준
        - name_col: 이름 열 번호
        - data_start: 데이터 시작 행 번호
    """
    # ── 헤더/날짜 열 찾기 ──
    header_row = None
    raw_day_cols = []  # [(col_index, calendar_day)]

    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            val = str(cell.value).strip() if cell.value else ""
            raw = val[:-1].strip() if val.endswith("일") else val
            try:
                d = int(raw)
                if 1 <= d <= 31:
                    if header_row is None:
                        header_row = cell.row
                    if cell.row == header_row:
                        raw_day_cols.append((cell.column, d))
            except ValueError:
                pass

    if not header_row or not raw_day_cols:
        return None

    # ── 열 순서대로 스케줄 day 1, 2, 3, ... 매핑 ──
    raw_day_cols.sort(key=lambda x: x[0])  # 열 위치 순 정렬
    day_cols = {i + 1: col for i, (col, _cal_day) in enumerate(raw_day_cols)}

    if not day_cols:
        return None

    # ── 이름 열 찾기 ──
    min_day_col = min(day_cols.values())
    name_col = 1  # 기본 A열

    for search_row in range(max(1, header_row - 1), header_row + 3):
        for row in ws.iter_rows(min_row=search_row, max_row=search_row,
                                max_col=min_day_col - 1):
            for cell in row:
                val = str(cell.value).strip() if cell.value else ""
                if val == "이름":
                    name_col = cell.column

    # ── 데이터 시작 행 찾기 ──
    data_start = header_row + 1
    for check in ws.iter_rows(min_row=header_row + 1,
                              max_row=min(header_row + 3, ws.max_row),
                              max_col=min_day_col + 6):
        for cell in check:
            val = str(cell.value).strip() if cell.value else ""
            if val in ("이름", "일", "월", "화", "수", "목", "금", "토"):
                data_start = check[0].row + 1
                break
        else:
            continue
        break

    return header_row, day_cols, name_col, data_start


def import_requests(
    filepath: str,
    nurses: list[Nurse],
    start_date: date,
    password: str | None = None,
) -> tuple[list[Request], dict[int, int]]:
    """근무신청표 엑셀에서 요청사항 + 간호사 속성 읽기

    Args:
        filepath: 엑셀 파일 경로
        nurses: 간호사 목록 (이름 매칭용, 속성도 업데이트됨)
        start_date: 스케줄 시작일

    Returns:
        (requests, weekly_off_map)
        - requests: Request 목록
        - weekly_off_map: {nurse_id: weekday_index} 고정 주휴 요일

    Side effects:
        nurses 리스트의 각 Nurse 객체에 아래 속성을 업데이트:
        - vacation_days: B열 (휴가 잔여일)
        - menstrual_used: C열 (생휴 이미 사용 여부)
        - pending_sleep: D열 (전월 수면 이월)

    형식 (응급실 실제):
      A열: 이름
      B열: 휴가 (잔여 연차)
      C열: 생휴 (1=이미 사용)
      D열: 수면 (값 있으면 전월 이월)
      E~AF열: 1일~28일 (코드 입력)
    """
    wb = load_workbook_safe(filepath, password, read_only=True, data_only=True)
    ws = wb.active

    nurse_name_map = {n.name.strip(): n for n in nurses}
    num_days = 28

    result = _find_day_columns(ws)
    if result is None:
        wb.close()
        return [], {}

    header_row, day_cols, name_col, data_start = result
    min_day_col = min(day_cols.values())

    # ── 휴가/생휴/수면 열 찾기 ──
    # 날짜 열 이전 컬럼에서만 검색 (뒤쪽 통계 영역 제외)
    vac_col = None
    menst_col = None
    sleep_col = None

    for search_row in range(max(1, header_row - 1), header_row + 3):
        for row in ws.iter_rows(min_row=search_row, max_row=search_row,
                                max_col=min_day_col - 1):
            for cell in row:
                val = str(cell.value).strip() if cell.value else ""
                if val in ("휴가", "연차"):
                    vac_col = cell.column
                elif val in ("생휴", "생리"):
                    menst_col = cell.column
                elif val == "수면":
                    sleep_col = cell.column

    # ── 데이터 읽기 ──
    requests = []
    weekly_off_map = {}
    stop_words = {"off", "주", "수면", "생휴", "vac", "공가", "총",
                  "d 인원", "중2 인원", "e 인원", "n 인원", "요일"}

    nid_counter = max([n.id for n in nurses], default=0) + 1 # 새로운 ID 시작점
    
    for row in ws.iter_rows(min_row=data_start):
        # 이름
        name_cell = row[name_col - 1] if name_col - 1 < len(row) else None
        name = str(name_cell.value).strip() if name_cell and name_cell.value else ""
        if not name:
            continue
        if name.lower() in stop_words or "인원" in name:
            break  # 집계 행 도달 → 종료
        if name not in nurse_name_map:
            new_nurse = Nurse(id=nid_counter, name=name)
            nurses.append(new_nurse)        # 원본 리스트에 추가
            nurse_name_map[name] = new_nurse # 맵에 추가
            nid_counter += 1

        nurse = nurse_name_map[name]
        nid = nurse.id

        # ── B열: 휴가 잔여일 ──
        if vac_col and vac_col - 1 < len(row):
            v = row[vac_col - 1].value
            if v is not None:
                try:
                    nurse.vacation_days = int(v)
                except (ValueError, TypeError):
                    pass

        # ── C열: 생휴 사용 여부 ──
        if menst_col and menst_col - 1 < len(row):
            v = row[menst_col - 1].value
            if v is not None:
                try:
                    nurse.menstrual_used = int(v) >= 1
                except (ValueError, TypeError):
                    nurse.menstrual_used = bool(v)

        # ── D열: 수면 이월 ──
        if sleep_col and sleep_col - 1 < len(row):
            v = row[sleep_col - 1].value
            if v is not None:
                val_str = str(v).strip()
                # "1", "1,2월 수면", 숫자 → 이월 있음
                nurse.pending_sleep = bool(val_str)

        # ── 날짜별 요청 ──
        first_weekly = None

        for d, col in day_cols.items():
            if d > num_days:
                continue
            cell = row[col - 1] if col - 1 < len(row) else None
            val = str(cell.value).strip() if cell and cell.value else ""
            if not val:
                continue

            # 슬래시 코드 → OR 요청 2건 생성
            if "/" in val:
                parts = val.split("/")
                or_codes = []
                for part in parts:
                    c = _normalize_code(part.strip())
                    if c is not None:
                        or_codes.append(c)
                for c in or_codes:
                    requests.append(Request(nurse_id=nid, day=d, code=c, is_or=True))
                continue

            code = _normalize_code(val)

            if code is None:
                continue

            # "주" → 고정 주휴 요일 감지
            if code == "주":
                wd = (start_date + timedelta(days=d - 1)).weekday()
                if first_weekly is None:
                    first_weekly = wd

            requests.append(Request(nurse_id=nid, day=d, code=code))

        if first_weekly is not None:
            weekly_off_map[nid] = first_weekly

    wb.close()
    return requests, weekly_off_map


def import_nurses_from_request(filepath: str, password: str | None = None) -> list[str]:
    """근무신청표에서 간호사 이름 목록만 추출

    Returns: 이름 리스트 (순서 유지)
    """
    wb = load_workbook_safe(filepath, password, read_only=True, data_only=True)
    ws = wb.active

    result = _find_day_columns(ws)
    if result is None:
        wb.close()
        return []

    _header_row, _day_cols, name_col, data_start = result

    # ── 이름 추출 ──
    names = []
    stop_words = {"off", "주", "수면", "생휴", "vac", "공가", "총",
                  "d 인원", "중2 인원", "e 인원", "n 인원", "요일"}

    for row in ws.iter_rows(min_row=data_start):
        cell = row[name_col - 1] if name_col - 1 < len(row) else None
        val = str(cell.value).strip() if cell and cell.value else ""
        if not val:
            continue
        if val.lower() in stop_words or "인원" in val:
            break  # 집계 행 도달
        names.append(val)

    wb.close()
    return names


# ══════════════════════════════════════════
# 가져오기: 이전 달 근무표 → 마지막 5일 근무
# ══════════════════════════════════════════

def import_prev_schedule(
    filepath: str,
    nurse_names: list[str],
    tail_days: int = 5,
    password: str | None = None,
) -> tuple[dict[str, list[str]], dict[str, int], dict[str, int], dict[str, int]]:
    """이전 달 근무표 엑셀에서 마지막 tail_days일의 근무 + 전체 N 횟수 + 수면 횟수 + 휴가잔여 추출

    export_schedule 형식 기준:
      행3: 헤더 (이름, 1~28, 통계...)
      행5+: 간호사별 (A열: 이름, B~AC열: 근무)

    Args:
        filepath: 엑셀 파일 경로
        nurse_names: 매칭할 간호사 이름 리스트
        tail_days: 추출할 마지막 일수 (기본 5)

    Returns:
        (tail_shifts, n_counts, sleep_counts, vac_days)
        - tail_shifts: {이름: [근무코드 리스트]} (가장 오래된 순)
        - n_counts: {이름: 전월 N 총 횟수}
        - sleep_counts: {이름: 전월 수면 사용 횟수}
        - vac_days: {이름: 휴가잔여 일수}
    """
    wb = load_workbook_safe(filepath, password, read_only=True, data_only=True)
    ws = wb.active

    result = _find_day_columns(ws)
    if result is None:
        wb.close()
        return {}, {}, {}

    header_row, day_cols, name_col, data_start = result

    max_day = max(day_cols.keys())
    all_day_range = sorted(day_cols.keys())
    tail_start = max(1, max_day - tail_days + 1)
    tail_day_range = [d for d in range(tail_start, max_day + 1) if d in day_cols]

    # 통계 열에서 "휴가잔여" 컬럼 위치 찾기 (날짜 열 이후)
    max_day_col = max(day_cols.values())
    vac_remain_col = None
    for row in ws.iter_rows(min_row=header_row, max_row=header_row):
        for col_idx, cell in enumerate(row, 1):
            if not hasattr(cell, 'value') or col_idx <= max_day_col:
                continue
            val = str(cell.value).strip() if cell.value else ""
            if val in ("휴가잔여", "휴가", "잔여", "잔여휴가", "연차잔여", "연차"):
                vac_remain_col = col_idx
                break
        if vac_remain_col:
            break

    name_set = set(n.strip() for n in nurse_names)
    tail_result = {}
    n_counts = {}
    sleep_counts = {}
    vac_days = {}
    stop_words = {"요일", "이름", "일", "월", "화", "수", "목", "금", "토",
                  "off", "주", "수면", "생휴", "vac", "공가", "총",
                  "d 인원", "중2 인원", "e 인원", "n 인원"}

    for row in ws.iter_rows(min_row=data_start):
        cell = row[name_col - 1] if name_col - 1 < len(row) else None
        name = str(cell.value).strip() if cell and cell.value else ""
        if not name or name.lower() in stop_words:
            continue
        if name not in name_set:
            continue

        # 마지막 tail_days일 근무
        shifts = []
        for d in tail_day_range:
            col = day_cols[d]
            c = row[col - 1] if col - 1 < len(row) else None
            val = str(c.value).strip() if c and c.value else ""
            code = _normalize_code(val) if val else ""
            shifts.append(code if code else val)
        tail_result[name] = shifts

        # 전체 N 횟수 + 수면 횟수
        n_count = 0
        sleep_count = 0
        for d in all_day_range:
            col = day_cols[d]
            c = row[col - 1] if col - 1 < len(row) else None
            val = str(c.value).strip() if c and c.value else ""
            if val == "N":
                n_count += 1
            if val == "수면":
                sleep_count += 1
        n_counts[name] = n_count
        sleep_counts[name] = sleep_count

        # 휴가잔여 추출
        if vac_remain_col and vac_remain_col - 1 < len(row):
            cell = row[vac_remain_col - 1]
            v = cell.value if hasattr(cell, 'value') else None
            if v is not None:
                try:
                    vac_days[name] = int(v)
                except (ValueError, TypeError):
                    pass

    wb.close()
    return tail_result, n_counts, sleep_counts, vac_days


# ══════════════════════════════════════════
# 파일 날짜 탐지
# ══════════════════════════════════════════

def detect_file_month(filepath: str, password: str | None = None) -> tuple[int | None, int | None]:
    """엑셀 파일에서 연도/월 정보를 탐지

    파일명 → 파일 내용 순으로 검색.
    고신뢰 패턴(YYYY.MM, YYYY년 M월)을 먼저 시도하고,
    없으면 짧은 셀에서 "X월" 패턴을 탐색.

    Returns:
        (year, month) — 탐지 실패 시 각각 None
    """
    import os

    # ── 1단계: 파일명에서 탐지 ──
    filename = os.path.basename(filepath)
    m = re.search(r'(\d{4})\s*[년._\-]\s*(\d{1,2})', filename)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        if 1 <= mo <= 12 and 2000 <= y <= 2099:
            return y, mo

    # ── 2단계: 파일 내용에서 탐지 ──
    wb = load_workbook_safe(filepath, password, read_only=True, data_only=True)
    ws = wb.active

    # Pass 1: 고신뢰 — "YYYY년 M월" 또는 "YYYY.MM"
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            val = str(cell.value).strip() if cell.value else ""
            if not val:
                continue

            # "2026년 3월" or "2026년3월"
            m = re.search(r'(\d{4})\s*년\s*(\d{1,2})\s*월', val)
            if m:
                wb.close()
                return int(m.group(1)), int(m.group(2))

            # "2026.03.01" or "2026-03-01" or "2026/3"
            m = re.search(r'(\d{4})[.\-/](\d{1,2})', val)
            if m:
                y, mo = int(m.group(1)), int(m.group(2))
                if 1 <= mo <= 12 and 2000 <= y <= 2099:
                    wb.close()
                    return y, mo

    # Pass 2: 저신뢰 — 짧은 셀(제목 등)에서 "X월" 탐색
    #   긴 안내 텍스트("9월 이상 야근..." 등)에서 오탐지 방지
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            val = str(cell.value).strip() if cell.value else ""
            if not val or len(val) > 30:
                continue
            m = re.search(r'(\d{1,2})\s*월', val)
            if m:
                mo = int(m.group(1))
                if 1 <= mo <= 12:
                    wb.close()
                    return None, mo

    wb.close()
    return None, None
