"""엑셀 가져오기 / 내보내기

내보내기: 근무표 + 통계 시트 → .xlsx
불러오기: 엑셀에서 간호사 목록, 요청사항 읽기
"""
import calendar
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from engine.models import Nurse, Request, Rules, Schedule


# 근무별 색상 (RGB hex)
FILLS = {
    "D":   PatternFill(start_color="DAF0F3", end_color="DAF0F3", fill_type="solid"),
    "E":   PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid"),
    "N":   PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"),
    "OFF": PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid"),
}
FONTS = {
    "D":   Font(color="2E75B6", bold=True, size=9),
    "E":   Font(color="C55A11", bold=True, size=9),
    "N":   Font(color="7030A0", bold=True, size=9),
    "OFF": Font(color="548235", bold=True, size=9),
}
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
WEEKEND_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
SHORTAGE_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
CENTER = Alignment(horizontal="center", vertical="center")


def export_schedule(schedule: Schedule, rules: Rules, filepath: str):
    """근무표 + 통계를 엑셀로 내보내기"""
    wb = Workbook()

    # ── Sheet 1: 근무표 ──
    ws = wb.active
    ws.title = "근무표"

    year, month = schedule.year, schedule.month
    num_days = schedule.num_days
    nurses = schedule.nurses
    weekday_names = ["월", "화", "수", "목", "금", "토", "일"]

    # 타이틀
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_days + 5)
    ws.cell(1, 1, f"{year}년 {month}월 간호사 근무표")
    ws.cell(1, 1).font = Font(bold=True, size=14, color="2F5496")
    ws.cell(1, 1).alignment = Alignment(horizontal="center")

    # 헤더 (행3)
    row = 3
    headers = ["이름"] + [f"{d}" for d in range(1, num_days + 1)] + ["D", "E", "N", "OFF"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # 요일 행 (행4)
    row = 4
    ws.cell(row, 1, "요일")
    ws.cell(row, 1).font = Font(bold=True, size=9)
    ws.cell(row, 1).alignment = CENTER
    for d in range(1, num_days + 1):
        wd = calendar.weekday(year, month, d)
        cell = ws.cell(row, d + 1, weekday_names[wd])
        cell.alignment = CENTER
        cell.font = Font(size=8, bold=True, color="CC0000" if wd >= 5 else "333333")
        if wd >= 5:
            cell.fill = WEEKEND_FILL

    # 간호사별 데이터
    for i, nurse in enumerate(nurses):
        row = 5 + i
        ws.cell(row, 1, nurse.name)
        ws.cell(row, 1).font = Font(bold=True, size=10)
        ws.cell(row, 1).alignment = CENTER
        ws.cell(row, 1).border = THIN_BORDER

        counts = {"D": 0, "E": 0, "N": 0, "OFF": 0}

        for d in range(1, num_days + 1):
            shift = schedule.get_shift(nurse.id, d)
            cell = ws.cell(row, d + 1, shift)
            cell.alignment = CENTER
            cell.border = THIN_BORDER

            if shift in FILLS:
                cell.fill = FILLS[shift]
                cell.font = FONTS[shift]
            else:
                wd = calendar.weekday(year, month, d)
                if wd >= 5:
                    cell.fill = WEEKEND_FILL

            if shift in counts:
                counts[shift] += 1

        # 통계 열
        for j, s in enumerate(["D", "E", "N", "OFF"]):
            cell = ws.cell(row, num_days + 2 + j, counts[s])
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            cell.font = Font(bold=True, size=9)

    # 빈 행
    sep_row = 5 + len(nurses)

    # 집계 행 (D/E/N 인원수)
    for si, shift_type in enumerate(["D", "E", "N"]):
        agg_row = sep_row + 1 + si
        ws.cell(agg_row, 1, f"{shift_type} 인원")
        ws.cell(agg_row, 1).font = Font(bold=True, size=9)
        ws.cell(agg_row, 1).alignment = CENTER

        for d in range(1, num_days + 1):
            count = schedule.get_staff_count(d, shift_type)
            cell = ws.cell(agg_row, d + 1, count)
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            cell.font = Font(bold=True, size=9)

            is_weekend = schedule.is_weekend(d)
            min_req = rules.get_min_staff(shift_type, is_weekend)
            if count < min_req:
                cell.fill = SHORTAGE_FILL
                cell.font = Font(bold=True, size=9, color="CC0000")

    # 컬럼 너비
    ws.column_dimensions["A"].width = 12
    for d in range(1, num_days + 1):
        ws.column_dimensions[get_column_letter(d + 1)].width = 5
    for j in range(4):
        ws.column_dimensions[get_column_letter(num_days + 2 + j)].width = 6

    # ── Sheet 2: 개인별 통계 ──
    ws2 = wb.create_sheet("통계")
    ws2.cell(1, 1, f"{year}년 {month}월 개인별 통계")
    ws2.cell(1, 1).font = Font(bold=True, size=14, color="2F5496")

    stat_headers = ["이름", "숙련도", "Day", "Evening", "Night", "OFF",
                     "총근무", "야간비율", "주말근무"]
    for c, h in enumerate(stat_headers, 1):
        cell = ws2.cell(3, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    weekend_days = [d for d in range(1, num_days + 1) if schedule.is_weekend(d)]

    for i, nurse in enumerate(nurses):
        row = 4 + i
        d_cnt = schedule.get_day_count(nurse.id, "D")
        e_cnt = schedule.get_day_count(nurse.id, "E")
        n_cnt = schedule.get_day_count(nurse.id, "N")
        off_cnt = schedule.get_day_count(nurse.id, "OFF")
        total_work = d_cnt + e_cnt + n_cnt
        night_ratio = f"{n_cnt/total_work*100:.0f}%" if total_work > 0 else "0%"
        weekend_work = sum(1 for d in weekend_days if schedule.get_shift(nurse.id, d) != "OFF")

        data = [nurse.name, nurse.skill_level, d_cnt, e_cnt, n_cnt, off_cnt,
                total_work, night_ratio, weekend_work]
        for c, val in enumerate(data, 1):
            cell = ws2.cell(row, c, val)
            cell.alignment = CENTER
            cell.border = THIN_BORDER

    for c in range(1, len(stat_headers) + 1):
        ws2.column_dimensions[get_column_letter(c)].width = 12

    wb.save(filepath)

def _detect_format(ws, header_row, headers):
    """엑셀 형식 감지: 'settings' (설정형) 또는 'calendar' (달력격자형)"""
    day_count = 0
    for cell in ws[header_row]:
        val = cell.value
        if val is not None:
            try:
                num = int(str(val).strip())
                if 1 <= num <= 31:
                    day_count += 1
            except (ValueError, TypeError):
                pass
    # 숫자 컬럼이 10개 이상이면 달력 격자
    if day_count >= 10:
        return "calendar"
    return "settings"

def import_nurses(filepath: str) -> list[Nurse]:
    """엑셀에서 간호사 목록 불러오기

    지원 형식 2가지:

    1) 설정 형식:
       이름 | 숙련도 | Day | Eve | Night | 비고

    2) 달력 격자 형식 (간호사들이 실제로 쓰는 형식):
       이름 | 1 | 2 | 3 | ... | 31
       김서연 | D |   | OFF |   | N
       → 이름만 추출, 나머지는 기본값
    """
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active

    # 헤더 찾기
    header_row = None
    headers = {}
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            val = str(cell.value).strip() if cell.value else ""
            if "이름" in val or "name" in val.lower():
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        return []

    # 헤더 매핑
    for cell in ws[header_row]:
        val = str(cell.value).strip().lower() if cell.value else ""
        col = cell.column
        if "이름" in val or "name" in val:
            headers["name"] = col
        elif "숙련" in val or "skill" in val or "경력" in val:
            headers["skill"] = col
        elif val in ("day", "d", "주간"):
            headers["can_day"] = col
        elif val in ("eve", "evening", "e", "저녁", "초번"):
            headers["can_evening"] = col
        elif val in ("night", "n", "야간", "밤"):
            headers["can_night"] = col
        elif "고정" in val or "fixed" in val:
            headers["fixed"] = col
        elif "비고" in val or "note" in val or "메모" in val:
            headers["note"] = col

    if "name" not in headers:
        wb.close()
        return []
    
    # 형식 감지
    fmt = _detect_format(ws, header_row, headers)

    if fmt == "calendar":
        # ── 달력 격자 형식: 이름만 추출 ──
        nurses = []
        nurse_id = 1
        for row in ws.iter_rows(min_row=header_row + 1):
            name_cell = row[headers["name"] - 1]
            name = str(name_cell.value).strip() if name_cell.vlaue else ""
            if not name:
                continue
            # D인원, E인원, N인원 같은 집계 행 제외
            if any(kw in name for kw in ["인원", "합계", "평균", "총"]):
                continue
            nurse = Nurse(id=nurse_id, name=name)
            nurses.append(nurse)
            nurse_id += 1
        wb.close()
        return nurses
    
    # ── 설정 형식: 기존 로직 ──
    nurses = []
    nurse_id = 1
    for row in ws.iter_rows(min_row=header_row + 1):
        name_cell = row[headers["name"] - 1]
        name = str(name_cell.value).strip() if name_cell.value else ""
        if not name:
            continue

        nurse = Nurse(id=nurse_id, name=name)

        if "skill" in headers:
            val = row[headers["skill"] - 1].value
            if val is not None:
                try:
                    nurse.skill_level = int(val)
                except (ValueError, TypeError):
                    pass

        for key, attr in [("can_day", "can_day"), ("can_evening", "can_evening"),
                          ("can_night", "can_night")]:
            if key in headers:
                val = str(row[headers[key] - 1].value).strip().upper() if row[headers[key] - 1].value else ""
                if val in ("X", "❌", "FALSE", "0", "불가", "N"):
                    setattr(nurse, attr, False)
                else:
                    setattr(nurse, attr, True)

        if "fixed" in headers:
            val = str(row[headers["fixed"] - 1].value).strip().upper() if row[headers["fixed"] - 1].value else ""
            if val in ("D", "E", "N"):
                nurse.fixed_shift = val

        if "note" in headers:
            val = row[headers["note"] - 1].value
            nurse.note = str(val).strip() if val else ""

        nurses.append(nurse)
        nurse_id += 1

    wb.close()
    return nurses


def import_requests(filepath: str, nurses: list[Nurse], year: int, month: int) -> list[Request]:
    """엑셀에서 요청사항 읽기 (달력 격자 형식)

    형식:
    이름 | 1 | 2 | 3 | ... | 31
    김서연 |   |OFF|   | ... | 연차
    """
    wb = load_workbook(filepath, read_only=True)

    # '요청' 시트 찾기 (없으면 첫 시트)
    ws = None
    for sheet_name in wb.sheetnames:
        if "요청" in sheet_name or "request" in sheet_name.lower():
            ws = wb[sheet_name]
            break
    if ws is None:
        ws = wb.active

    nurse_name_map = {n.name.strip(): n.id for n in nurses}
    num_days = calendar.monthrange(year, month)[1]
    valid_codes = {"OFF", "연차", "D", "E", "N", "D!", "E!", "N!"}
    requests = []

    # 헤더에서 날짜 열 찾기
    header_row = None
    day_cols = {}
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            val = str(cell.value).strip() if cell.value else ""
            if "이름" in val or "name" in val.lower():
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        wb.close()
        return []

    # 날짜 열 매핑
    for cell in ws[header_row]:
        val = cell.value
        if val is not None:
            try:
                d = int(str(val).strip())
                if 1 <= d <= num_days:
                    day_cols[d] = cell.column
            except (ValueError, TypeError):
                pass

    # 이름 열 찾기
    name_col = None
    for cell in ws[header_row]:
        val = str(cell.value).strip() if cell.value else ""
        if "이름" in val or "name" in val.lower():
            name_col = cell.column
            break

    if not name_col or not day_cols:
        wb.close()
        return []

    # 데이터 읽기
    for row in ws.iter_rows(min_row=header_row + 1):
        name = str(row[name_col - 1].value).strip() if row[name_col - 1].value else ""
        if name not in nurse_name_map:
            continue
        nurse_id = nurse_name_map[name]

        for d, col in day_cols.items():
            val = str(row[col - 1].value).strip() if row[col - 1].value else ""
            if val in valid_codes:
                requests.append(Request(nurse_id=nurse_id, day=d, code=val))

    wb.close()
    return requests
